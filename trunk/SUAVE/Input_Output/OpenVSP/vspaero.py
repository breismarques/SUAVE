#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Sat Oct 28 22:59:23 2017
@author: root
"""

import SUAVE
from SUAVE.Core import Units, Data
from subprocess import call

import xlwt
import string
import openpyxl
import io
import os

try:
    import sys
    sys.path.insert(0, '/Users/Bruno/OpenVSP/build/python_api')
    import vsp_g as vsp
    import vsp as vsp1

except ImportError:
    # This allows SUAVE to build without OpenVSP
    pass
import numpy as np


## @ingroup Input_Output-OpenVSP
def vspaero(tag,rho,vel,AoA,MachNumber,NumberIterations, rpm_forward, rpm_lift, engines_number_tot, Cp_lift, Cp_forward, Ct_lift, Ct_forward):
    
    if 1==1:

        try:
            
            vsp.ClearVSPModel()
            
        except NameError:
            print 'VSP import failed'
            
        vsp.VSPCheckSetup()
        vsp.VSPRenew()


        #open the file created in vsp_write
    
        vsp.ReadVSPFile(tag)
        
        #==== Analysis: VSPAero Compute Geometry to Create Vortex Lattice DegenGeom File ====//
        
        compgeom_name = "VSPAEROComputeGeometry";

        # Set defaults
        vsp.SetAnalysisInputDefaults(compgeom_name);

        # list inputs, type, and current values
        vsp.PrintAnalysisInputs(compgeom_name);

        # Execute
        compgeom_resid=vsp.ExecAnalysis(compgeom_name);

        # Get & Display Results
        vsp.PrintResults( compgeom_resid );
        
    
        #==== Analysis: VSPAero Compute Geometry ====//
    
        analysis_name="VSPAEROSweep"
    
        #Set defaults
    
        vsp.SetAnalysisInputDefaults(analysis_name)
    
        #Change some input values
        #    Analysis method
    
        analysis_method = [vsp.VORTEX_LATTICE]
    
    
        vsp.SetIntAnalysisInput( analysis_name, "AnalysisMethod", analysis_method, 0 )
        
        #Reference geometry set
        
        geom_set=[0]
        vsp.SetIntAnalysisInput( analysis_name, "GeomSet", geom_set );
                               
        #Reference areas, lengths
        
        wing_id=vsp.FindGeomsWithName("main_wing")
        
        sref=[float(0)]*1
        bref=[float(0)]*1
        cref=[float(0)]*1
        
        
        sref[0]=(vsp.GetParmVal(wing_id[0],"TotalArea", "WingGeom"))
        bref[0]=(vsp.GetParmVal(wing_id[0],"TotalSpan", "WingGeom"))
        cref[0]=(vsp.GetParmVal(wing_id[0],"TotalChord", "WingGeom"))
        
        print sref
        print bref
        print cref
        
        ref_flag=[3]
        
        
        vsp.SetDoubleAnalysisInput( analysis_name, 'Sref', sref )
        vsp.SetDoubleAnalysisInput( analysis_name, 'bref', bref )
        vsp.SetDoubleAnalysisInput( analysis_name, 'cref', cref )
        vsp.SetIntAnalysisInput( analysis_name, "RefFlag", ref_flag )
        
                                 
        #Freestream parameters
        #Alpha
        
        alpha_start=[float(AoA)]
        alpha_end=[float(AoA)]
        alpha_npts=[1]
        
        print alpha_start
        
        vsp.SetDoubleAnalysisInput( analysis_name, "AlphaStart", alpha_start )
        vsp.SetDoubleAnalysisInput( analysis_name, "AlphaEnd", alpha_end )
        vsp.SetIntAnalysisInput( analysis_name, "AlphaNpts", alpha_npts )
        
        #Beta
        beta_start=[float(0)]
        beta_end=[float(0)]
        beta_npts=[1]
        
        
        vsp.SetDoubleAnalysisInput( analysis_name, "BetaStart", beta_start )
        vsp.SetDoubleAnalysisInput( analysis_name, "BetaEnd", beta_end )
        vsp.SetIntAnalysisInput( analysis_name, "BetaNpts", beta_npts );
                               
        #Mach
        
        mach_start=[float(MachNumber)]
        mach_end=[float(MachNumber)]
        mach_npts=[1]
        
        print mach_start
        
        vsp.SetDoubleAnalysisInput( analysis_name, "MachStart", mach_start )
        vsp.SetDoubleAnalysisInput( analysis_name, "MachEnd", mach_end )
        vsp.SetIntAnalysisInput( analysis_name, "MachNpts", mach_npts )
                                
        vsp.Update()
        
        #Vinf
        
        vel_rotor=float(vel)
        print vel_rotor
        
        vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
        vel_id = vsp.FindParm( vspaero_settings_container_id, 'Vinf', 'VSPAERO')
        vsp.SetParmVal( vel_id, vel_rotor)
        
        #rho
        
        rho_rotor=float(rho)
        print rho_rotor
        
        vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
        rho_id = vsp.FindParm( vspaero_settings_container_id, 'Rho', 'VSPAERO')
        vsp.SetParmVal( rho_id, rho_rotor)
        
        
        #rpm
        
        number_eng = int(engines_number_tot)
        print number_eng
        forward_rpm=float(rpm_forward)
        print forward_rpm
        lift_rpm=float(rpm_lift)
        print lift_rpm
        
        g=int(0)
        
        while g<number_eng:
            
            if g<2:
                
                if forward_rpm==0.0:
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    rpm_id = vsp.FindParm( vspaero_settings_container_id, 'RotorRPM', 'Rotor_'+str(g))
                    vsp.SetParmVal( rpm_id, float(0.001))
                else:
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    rpm_id = vsp.FindParm( vspaero_settings_container_id, 'RotorRPM', 'Rotor_'+str(g))
                    vsp.SetParmVal( rpm_id, forward_rpm)

            else:
                
                if lift_rpm==0.0:
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    rpm_id = vsp.FindParm( vspaero_settings_container_id, 'RotorRPM', 'Rotor_'+str(g))
                    vsp.SetParmVal( rpm_id, float(0.001))
                else:
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    rpm_id = vsp.FindParm( vspaero_settings_container_id, 'RotorRPM', 'Rotor_'+str(g))
                    vsp.SetParmVal( rpm_id, lift_rpm)

            g=g+1
            
        #vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
        #rpm_id = vsp.FindParm( vspaero_settings_container_id, 'RotorRPM', 'Rotor_0')
        #vsp.SetParmVal( rpm_id, 3000.0)
            
        #Power Coefficient
        
        forward_Cp=float(Cp_forward)
        print forward_Cp
        lift_Cp=float(Cp_lift)
        print lift_Cp
        
        
        g=int(0)
        
        while g<number_eng:
            
            if g<2:
                
                if forward_Cp<-1000.00000:
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    cp_id = vsp.FindParm( vspaero_settings_container_id, 'RotorCP', 'Rotor_'+str(g))
                    vsp.SetParmVal( cp_id, -1000.00000)
                    
                elif forward_Cp>1000.00000:
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    cp_id = vsp.FindParm( vspaero_settings_container_id, 'RotorCP', 'Rotor_'+str(g))
                    vsp.SetParmVal( cp_id, 1000.00000)
                    
                else:
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    cp_id = vsp.FindParm( vspaero_settings_container_id, 'RotorCP', 'Rotor_'+str(g))
                    vsp.SetParmVal( cp_id, forward_Cp)
                
            

            else:
                
                if lift_Cp<-1000.00000:
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    cp_id = vsp.FindParm( vspaero_settings_container_id, 'RotorCP', 'Rotor_'+str(g))
                    vsp.SetParmVal( cp_id, -1000.00000)
                    
                elif lift_Cp>1000.00000:
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    cp_id = vsp.FindParm( vspaero_settings_container_id, 'RotorCP', 'Rotor_'+str(g))
                    vsp.SetParmVal( cp_id, 1000.00000)
                    
                else:
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    cp_id = vsp.FindParm( vspaero_settings_container_id, 'RotorCP', 'Rotor_'+str(g))
                    vsp.SetParmVal( cp_id, lift_Cp)

            g=g+1
            
        
        #Thrust Coefficient
        
        forward_Ct=float(Ct_forward)
        print forward_Ct
        lift_Ct=float(Ct_lift)
        print lift_Ct
        
        
        g=int(0)
        
        while g<number_eng:
            
            if g<2:
                
                if 0.001 < forward_Ct < 1000.00000:
                    
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    ct_id = vsp.FindParm( vspaero_settings_container_id, 'RotorCT', 'Rotor_'+str(g))
                    vsp.SetParmVal( ct_id, forward_Ct)
                    
                elif forward_Ct > 1000.00000:
                    
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    ct_id = vsp.FindParm( vspaero_settings_container_id, 'RotorCT', 'Rotor_'+str(g))
                    vsp.SetParmVal( ct_id, 1000.00000)
                    
                else:
                    
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    ct_id = vsp.FindParm( vspaero_settings_container_id, 'RotorCT', 'Rotor_'+str(g))
                    vsp.SetParmVal( ct_id, 0.001)
                               

            else:
                
                if  0.001<lift_Ct<1000.00000:
                    
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    ct_id = vsp.FindParm( vspaero_settings_container_id, 'RotorCT', 'Rotor_'+str(g))
                    vsp.SetParmVal( ct_id, lift_Ct)
                    
                elif lift_Ct>1000.00000:
                    
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    ct_id = vsp.FindParm( vspaero_settings_container_id, 'RotorCT', 'Rotor_'+str(g))
                    vsp.SetParmVal( ct_id, 1000.00000)
                    
                else:
                    
                    vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
                    ct_id = vsp.FindParm( vspaero_settings_container_id, 'RotorCT', 'Rotor_'+str(g))
                    vsp.SetParmVal( ct_id, 0.001)
                              

            g=g+1
                
                
        
        #vsp.SetDoubleAnalysisInput( analysis_name, "MachABFDS", mach_start )
        
        #vspaero_settings_container_id = vsp.FindContainer( "VSPAEROSettings", 0 )
        #rpm_cruise_id = vsp.FindParm( vspaero_settings_container_id, 'RotorRPM', 'Rotor_0')
        #vsp.SetParmValUpdate( rpm_cruise_id, 1000.0)
                                
        #vsp.Update()
                  
        ##Case Setup
        
        wakeNumIter=[int(NumberIterations)]
        wakeSkipUntilIter=[int(NumberIterations+1)]
        batch_mode_flag=[1]
        
        vsp.SetIntAnalysisInput( analysis_name, "WakeNumIter", wakeNumIter )
        vsp.SetIntAnalysisInput( analysis_name, "WakeSkipUntilIter", wakeSkipUntilIter )
        vsp.SetIntAnalysisInput( analysis_name, "BatchModeFlag", batch_mode_flag )
                                
        vsp.Update()
    
    
        #list inputs, type, and current values
    
        vsp.PrintAnalysisInputs(analysis_name)
        
    
        #Execute
    
        rid = vsp.ExecAnalysis(analysis_name)
            
    
        #Get & Display Results
    
        vsp.PrintResults(rid)
                        
        #Write in CSV
        
        csvname='Simulation_'+tag[:-5]
        
        vsp.WriteResultsCSVFile(rid,csvname+'.csv')
        
        #Close File
        
        vsp.ClearVSPModel();
        vsp.Update();
    
        # Check for errors

        errorMgr = vsp.ErrorMgrSingleton_getInstance()
        num_err = errorMgr.GetNumTotalErrors()
        for i in range(0, num_err):
            err = errorMgr.PopLastError()
            print("error = ", err.m_ErrorString)
            
        print 'FINISHED VSPAERO SIMULATION'
        
        data = []
        with open(csvname+'.csv') as f:
            for line in f:
                data.append([word for word in line.split(",") if word])
        f.close()
        
        myfile="/Users/Bruno/Documents/Delft/Courses/2016-2017/Thesis/Code/Bruno_Aircraft/Optimization_Lo_Fid/"+csvname+'.csv'

        ## If file exists, delete it ##
        if os.path.isfile(myfile):
            os.remove(myfile)
        else:    ## Show an error ##
            print("Error: %s file not found" % myfile)
        
        
        
        #print float(lift[:-1])
        #print float(drag[:-1])
        
        try:
            
            lift=data[18][NumberIterations]
            
        except IndexError:
            
            print 'INDEX ERROR'
            lift='-10.0000'
            
        try:
            
            drag=data[14][NumberIterations]
            
        except IndexError:
            
            print 'INDEX ERROR'
            drag='-10.0000'
            
                
        #wb = openpyxl.Workbook()
        #sheet = wb.active
        #for row_index in range(len(data)):
        #    for col_index, letter in zip(range(len(data[row_index])), string.ascii_uppercase):
        #        sheet[letter+str(row_index+1)]= data[row_index][col_index]

        #wb.save(csvname+'.xlsx')
        
        #xlsx_filename=csvname+'.xlsx'
        
        #with open(xlsx_filename, "rb") as f:
        #    in_mem_file = io.BytesIO(f.read())

        #book = openpyxl.load_workbook(in_mem_file, read_only=True)
        
        #book = openpyxl.load_workbook(csvname+'.xlsx')
        
        #sheet = book.active
        
        #CL = sheet[chr(NumberIterations+65)+'19'].value
        #CD = sheet[chr(NumberIterations+65)+'15'].value
                   
        CL = float(lift[:-1])
        CD = float(drag[:-1])
                   
        #book.save(csvname+'.xlsx')
                   
        #call(["cd","/Users/Bruno/Documents/Delft/Courses/2016-2017/Thesis/Code/Bruno_Aircraft","&&","echo","iforgot","|","sudo","-S","rm","Simulation_takeoff.xlsx"])

        #call(["echo","iforgot","|","sudo","-S","rm","Simulatio.xlsx"])            
        
        
    
    return CL, CD